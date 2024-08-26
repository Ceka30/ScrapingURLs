import requests
from bs4 import BeautifulSoup
from PIL import Image, UnidentifiedImageError
from io import BytesIO
import pandas as pd
import os
from urllib.parse import urljoin
from selenium import webdriver
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

def get_image_urls(page_url):
    try:
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        driver = webdriver.Chrome(options=options)
        driver.get(page_url)

        soup = BeautifulSoup(driver.page_source, 'html.parser')
        driver.quit()

        img_tags = soup.find_all('img')
        img_urls = [urljoin(page_url, img['src']) for img in img_tags if 'src' in img.attrs]
        return img_urls
    except Exception as e:
        print(f"Error al obtener imágenes de {page_url}: {e}")
        return []

def get_image_size(url):
    try:
        # Se ignora la verificación del certificado SSL al hacer la solicitud
        response = requests.get(url, stream=True, verify=False)
        response.raise_for_status()

        content_type = response.headers.get('Content-Type', '')

        if content_type.startswith('image/'):
            if 'svg' in content_type:
                size_kb = len(response.content) / 1024
                return round(size_kb, 1)

            img = Image.open(BytesIO(response.content))
            size_kb = len(response.content) / 1024
            return round(size_kb, 1)
        
        return None
    except UnidentifiedImageError:
        print(f"No se pudo identificar la imagen en {url}")
        return None
    except Exception as e:
        print(f"Error al obtener la imagen de {url}: {e}")
        return None

def process_pages(url_file, output_excel):
    urls = []
    with open(url_file, 'r') as file:
        urls = file.read().splitlines()

    data = []
    for page_url in urls:
        img_urls = get_image_urls(page_url)
        for img_url in img_urls:
            size_kb = get_image_size(img_url)
            if size_kb is not None:
                status = 'OK' if size_kb <= 500 else 'NO OK'
                data.append({'URL': page_url, 'Imagen': img_url, 'Tamaño (KB)': size_kb, 'Estado': status})

    df = pd.DataFrame(data)

    try:
        df.to_excel(output_excel, index=False)

        wb = load_workbook(output_excel)
        ws = wb.active

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        alignment = Alignment(horizontal="left", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
            for cell in row:
                if cell.column_letter == 'C':
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0'

        wb.save(output_excel)
        print("Análisis finalizado. Los resultados se han guardado en", output_excel)

    except PermissionError:
        print(f"Error: No se pudo escribir el archivo {output_excel} porque está en uso o no se tienen permisos.")
        return

url_file = 'urls.txt'
output_excel = 'resultados.xlsx'

if os.path.exists(output_excel):
    try:
        os.rename(output_excel, output_excel)
    except OSError:
        print(f"Error: El archivo {output_excel} está en uso. Por favor ciérrelo y vuelva a intentarlo.")
else:
    process_pages(url_file, output_excel)
